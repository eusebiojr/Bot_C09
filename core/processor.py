# core/processor.py
"""
Módulo de processamento de dados C09.
Substitui as funções tratar_planilha_c09() duplicadas em RRP/TLS.
VERSÃO CORRIGIDA: Com validações robustas para Cloud Run.
"""

import pandas as pd
import unicodedata
import os
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from typing import Dict, List, Tuple, Any


class C09DataProcessor:
    """
    Processador genérico para dados C09.
    Configurável via planilha de POIs e SLAs.
    """
    
    def __init__(self, config_pois: List[Dict[str, Any]]):
        """
        Inicializa processador com configuração de POIs.
        
        Args:
            config_pois: Lista de dicionários com configuração dos POIs
                        Formato: [{"ponto_interesse": str, "grupo": str, "sla_horas": float}, ...]
        """
        self.config_pois = config_pois
        self.pontos_desejados = [poi["ponto_interesse"] for poi in config_pois if poi.get("ativo", True)]
        self.mapa_grupos = {poi["ponto_interesse"]: poi["grupo"] for poi in config_pois}
        self.slas = self._calcular_slas()
        
    def _calcular_slas(self) -> Dict[str, float]:
        """Calcula SLAs com buffer de 30% por grupo."""
        slas = {}
        
        for poi in self.config_pois:
            if poi.get("sla_horas", 0) > 0:  # Só considera POIs com SLA definido
                grupo = poi["grupo"]
                sla_base = poi["sla_horas"]
                sla_com_buffer = sla_base * 1.3  # 30% de buffer
                
                # Mapeia para tipos de SLA reconhecidos
                if grupo == "Carregamento":
                    slas["carga"] = sla_com_buffer
                elif grupo == "Descarregamento":
                    slas["descarga"] = sla_com_buffer
                elif grupo == "Fabrica":
                    slas["carga"] = sla_com_buffer
                elif grupo == "Terminal":
                    slas["descarga"] = sla_com_buffer
        
        # SLAs de trajeto (valores padrão que podem ser configurados depois)
        slas.setdefault("trajeto_carregado", 6.3667 * 1.3)  # RRP padrão
        slas.setdefault("trajeto_vazio", 6.0833 * 1.3)      # RRP padrão
        
        return slas
    
    def _padronizar_texto(self, texto: str) -> str:
        """Remove acentos e padroniza texto para comparação."""
        if pd.isna(texto):
            return ""
        
        # Remove acentos e normaliza
        texto_normalizado = unicodedata.normalize("NFKD", str(texto))
        texto_ascii = texto_normalizado.encode("ascii", errors="ignore").decode("utf-8")
        
        return texto_ascii.strip()
    
    def _classificar_grupo(self, ponto: str) -> str:
        """Classifica grupo do POI com base na configuração."""
        return self.mapa_grupos.get(ponto, "Outros")
    
    def _agrupar_registros_consecutivos(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Agrupa registros consecutivos do mesmo veículo no mesmo POI.
        
        Args:
            df: DataFrame com dados filtrados
            
        Returns:
            DataFrame agrupado
        """
        # Aplica a classificação de grupo
        df["Grupo"] = df["Ponto de Interesse"].apply(self._classificar_grupo)

        # Agrupa entradas consecutivas por Veículo e Ponto de Interesse
        agrupados = []
        df_reset = df.reset_index(drop=True)
        idx = 0
        while idx < len(df_reset):
            atual = df_reset.iloc[idx]
            veic  = atual["Veículo"]
            ponto = atual["Ponto de Interesse"]
            grupo = atual["Grupo"]
            entrada = atual["Data Entrada"]
            saida   = atual["Data Saída"]

            j = idx + 1
            while j < len(df_reset) and df_reset.iloc[j]["Veículo"] == veic and df_reset.iloc[j]["Ponto de Interesse"] == ponto:
                saida = df_reset.iloc[j]["Data Saída"]
                j += 1

            agrupados.append({
                "Veículo": veic,
                "Ponto de Interesse": ponto,
                "Data Entrada": entrada,
                "Data Saída": saida,
                "Grupo": grupo,
                "Observações": atual.get("Observações", "")
            })
            idx = j

        df_ag = pd.DataFrame(agrupados)

        # Calcula o tempo de permanência
        df_ag["Tempo (h)"] = (df_ag["Data Saída"] - df_ag["Data Entrada"]).dt.total_seconds() / 3600
        df_ag["Tempo (h)"] = df_ag["Tempo (h)"].round(5)

        # Adiciona colunas necessárias
        df_ag["Trajeto Carregado"] = 0.0
        df_ag["Trajeto Vazio"] = 0.0
        df_ag["Observação"] = ""

        return df_ag
    
    def _calcular_trajetos(self, df_ag: pd.DataFrame) -> pd.DataFrame:
        """Calcula trajetos carregados e vazios."""
        def soma_justificativas(df_temp, i_inicio, i_fim):
            """Soma horas de manutenção e parada operacional entre dois índices."""
            horas_manutencao = 0.0
            horas_operacional = 0.0
            
            for k in range(i_inicio + 1, i_fim):
                registro = df_temp.iloc[k]
                if registro["Grupo"] == "Manutenção":
                    horas_manutencao += registro["Tempo (h)"]
                elif registro["Grupo"] == "Parada Operacional":
                    horas_operacional += registro["Tempo (h)"]
            
            return horas_manutencao, horas_operacional
        
        # Processa cada linha para calcular trajetos
        for i in range(len(df_ag)):
            atual = df_ag.iloc[i]
            veic = atual["Veículo"]
            grupo = atual["Grupo"]
            saida_atual = atual["Data Saída"]
            
            if grupo in ["Carregamento", "Fabrica"]:
                # Trajeto Carregado: desta saída até próxima entrada de Descarregamento/Terminal
                for j in range(i + 1, len(df_ag)):
                    prox = df_ag.iloc[j]
                    if prox["Veículo"] != veic:
                        break
                    if prox["Grupo"] in ["Descarregamento", "Terminal"]:
                        if prox["Data Entrada"] >= saida_atual:
                            delta_h = (prox["Data Entrada"] - saida_atual).total_seconds() / 3600
                            df_ag.at[i, "Trajeto Carregado"] = round(delta_h, 5)
                            
                            if delta_h > self.slas.get("trajeto_carregado", float('inf')):
                                horas_mant, horas_oper = soma_justificativas(df_ag, i, j)
                                if (horas_mant + horas_oper) > 0:
                                    df_ag.at[i, "Observação"] += (
                                        f"Trajeto Carregado longo ({delta_h:.2f}h > {self.slas['trajeto_carregado']:.2f}h): "
                                        f"{horas_mant:.2f}h em Manutenção, {horas_oper:.2f}h em Parada Operacional. "
                                    )
                                else:
                                    df_ag.at[i, "Observação"] += (
                                        f"Trajeto Carregado longo ({delta_h:.2f}h > {self.slas['trajeto_carregado']:.2f}h), sem justificativa. "
                                    )
                        break
                        
            elif grupo in ["Descarregamento", "Terminal"]:
                # Trajeto Vazio: desta saída até próxima entrada de Carregamento
                for j in range(i + 1, len(df_ag)):
                    prox = df_ag.iloc[j]
                    if prox["Veículo"] != veic:
                        break
                    if prox["Grupo"] in ["Carregamento", "Fabrica"]:
                        if prox["Data Entrada"] >= saida_atual:
                            delta_h = (prox["Data Entrada"] - saida_atual).total_seconds() / 3600
                            df_ag.at[i, "Trajeto Vazio"] = round(delta_h, 5)
                            
                            if delta_h > self.slas.get("trajeto_vazio", float('inf')):
                                horas_mant, horas_oper = soma_justificativas(df_ag, i, j)
                                if (horas_mant + horas_oper) > 0:
                                    df_ag.at[i, "Observação"] += (
                                        f"Trajeto Vazio longo ({delta_h:.2f}h > {self.slas['trajeto_vazio']:.2f}h): "
                                        f"{horas_mant:.2f}h em Manutenção, {horas_oper:.2f}h em Parada Operacional. "
                                    )
                                else:
                                    df_ag.at[i, "Observação"] += (
                                        f"Trajeto Vazio longo ({delta_h:.2f}h > {self.slas['trajeto_vazio']:.2f}h), sem justificativa. "
                                    )
                        break
        
        return df_ag
    
    def _formatar_dados_finais(self, df: pd.DataFrame) -> pd.DataFrame:
        """Aplica formatações finais aos dados."""
        df = df.copy()
        
        # Limpa observações
        df["Observação"] = df["Observação"].str.strip()
        
        # Formata veículo (remove prefixo se houver)
        df["Veículo"] = df["Veículo"].astype(str).apply(
            lambda x: x.split("-")[-1].strip()
        )
        
        # Formata datas para string
        df["Data Entrada"] = df["Data Entrada"].dt.strftime("%d/%m/%Y %H:%M:%S")
        df["Data Saída"] = df["Data Saída"].dt.strftime("%d/%m/%Y %H:%M:%S")
        
        return df
    
    def _criar_excel_formatado(self, df: pd.DataFrame) -> BytesIO:
        """Cria Excel formatado com tabela."""
        buffer_excel = BytesIO()
        
        # Escreve dados
        with pd.ExcelWriter(buffer_excel, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Relatório")
        
        buffer_excel.seek(0)
        
        # Formata como tabela
        wb = load_workbook(buffer_excel)
        ws = wb.active
        num_linhas = ws.max_row
        num_colunas = ws.max_column
        
        if num_linhas > 1 and num_colunas > 0:
            # Calcula referência da tabela
            def coluna_letra(n):
                letra = ""
                while n > 0:
                    n, resto = divmod(n - 1, 26)
                    letra = chr(65 + resto) + letra
                return letra
            
            ultima_coluna = coluna_letra(num_colunas)
            ref = f"A1:{ultima_coluna}{num_linhas}"
            
            # Cria tabela
            tabela = Table(displayName="TabelaRelatorio", ref=ref)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tabela.tableStyleInfo = style
            ws.add_table(tabela)
        
        # Salva resultado final
        final_buffer = BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)
        
        return final_buffer
    
    def processar_relatorio_c09(self, caminho_arquivo_origem: str) -> BytesIO:
        """
        Processa relatório C09 completo.
        VERSÃO CORRIGIDA: Com validações robustas.
        
        Args:
            caminho_arquivo_origem: Caminho do arquivo Excel baixado
            
        Returns:
            BytesIO: Excel tratado e formatado
            
        Raises:
            Exception: Se houver erro no processamento
        """
        try:
            print(f"=== Iniciando processamento: {caminho_arquivo_origem} ===")
            
            # ===== VALIDAÇÕES CRÍTICAS (ADICIONADAS) =====
            if not caminho_arquivo_origem:
                print("❌ ERRO: Caminho do arquivo é None ou vazio")
                raise ValueError("Caminho do arquivo é None ou vazio")
            
            if not os.path.exists(caminho_arquivo_origem):
                print(f"❌ ERRO: Arquivo não encontrado: {caminho_arquivo_origem}")
                raise FileNotFoundError(f"Arquivo não encontrado: {caminho_arquivo_origem}")
            
            tamanho_arquivo = os.path.getsize(caminho_arquivo_origem)
            if tamanho_arquivo == 0:
                print(f"❌ ERRO: Arquivo vazio: {caminho_arquivo_origem}")
                raise ValueError(f"Arquivo vazio: {caminho_arquivo_origem}")
            
            print(f"✅ Validação OK: {caminho_arquivo_origem} ({tamanho_arquivo} bytes)")
            # ===== FIM DAS VALIDAÇÕES =====
            
            # 1. Carrega dados
            df = pd.read_excel(caminho_arquivo_origem)
            print(f"Dados carregados: {len(df)} registros")
            
            if df.empty:
                raise ValueError("Arquivo Excel está vazio ou sem dados válidos")
            
            # 2. Padroniza POIs
            df["Ponto de Interesse"] = df["Ponto de Interesse"].astype(str).apply(self._padronizar_texto)
            
            # 3. Filtra POIs desejados
            df_filtrado = df[df["Ponto de Interesse"].isin(self.pontos_desejados)].copy()
            print(f"Após filtro POIs: {len(df_filtrado)} registros")
            
            if df_filtrado.empty:
                raise ValueError("Nenhum registro encontrado após filtro de POIs")
            
            # 4. Organiza dados
            df_filtrado = df_filtrado.sort_values(by=["Veículo", "Data Entrada"])
            df_filtrado = df_filtrado[["Veículo", "Ponto de Interesse", "Data Entrada", "Data Saída", "Observações"]]
            
            # 5. Agrupa registros consecutivos
            df_agrupado = self._agrupar_registros_consecutivos(df_filtrado)
            print(f"Após agrupamento: {len(df_agrupado)} registros")
            
            # 6. Calcula trajetos e SLAs
            df_com_trajetos = self._calcular_trajetos(df_agrupado)
            
            # 7. Formatação final
            df_final = self._formatar_dados_finais(df_com_trajetos)
            
            # 8. Cria Excel formatado
            buffer_resultado = self._criar_excel_formatado(df_final)
            
            print(f"=== Processamento concluído: {len(df_final)} registros finais ===")
            return buffer_resultado
            
        except Exception as e:
            print(f"❌ ERRO no processamento: {e}")
            print(f"   - Arquivo: {caminho_arquivo_origem}")
            print(f"   - Tipo do erro: {type(e).__name__}")
            raise


class ConfiguradorSLAs:
    """Classe para configurar SLAs específicos por unidade."""
    
    @staticmethod
    def aplicar_slas_rrp(config_pois: List[Dict]) -> List[Dict]:
        """Aplica SLAs específicos do RRP."""
        slas_rrp = {
            "trajeto_carregado": 6.3667,     # 6:22h
            "trajeto_vazio": 6.0833,         # 6:05h
        }
        
        for poi in config_pois:
            if poi["grupo"] == "Carregamento" and poi["ponto_interesse"] == "Carregamento RRp":
                poi["sla_horas"] = 1.0
            elif poi["grupo"] == "Descarregamento" and poi["ponto_interesse"] == "Descarga Inocencia":
                poi["sla_horas"] = 1.1833  # 1:11h
        
        return config_pois
    
    @staticmethod
    def aplicar_slas_tls(config_pois: List[Dict]) -> List[Dict]:
        """Aplica SLAs específicos do TLS."""
        slas_tls = {
            "trajeto_carregado": 3.5,     # 3:30h
            "trajeto_vazio": 2.9733,      # 2:58:24h
        }
        
        for poi in config_pois:
            if poi["grupo"] == "Carregamento" and poi["ponto_interesse"] == "Carregamento Fabrica":
                poi["sla_horas"] = 1.0
            elif poi["grupo"] == "Descarregamento" and poi["ponto_interesse"] == "Descarga TAP":
                poi["sla_horas"] = 0.9167  # 0:55h
        
        return config_pois


# Factory functions para compatibilidade
def criar_processor_rrp(config_pois: List[Dict]) -> C09DataProcessor:
    """Cria processador configurado para RRP."""
    config_ajustada = ConfiguradorSLAs.aplicar_slas_rrp(config_pois)
    processor = C09DataProcessor(config_ajustada)
    
    # Ajusta SLAs de trajeto específicos do RRP
    processor.slas["trajeto_carregado"] = 6.3667 * 1.3
    processor.slas["trajeto_vazio"] = 6.0833 * 1.3
    
    return processor


def criar_processor_tls(config_pois: List[Dict]) -> C09DataProcessor:
    """Cria processador configurado para TLS.""" 
    config_ajustada = ConfiguradorSLAs.aplicar_slas_tls(config_pois)
    processor = C09DataProcessor(config_ajustada)
    
    # Ajusta SLAs de trajeto específicos do TLS
    processor.slas["trajeto_carregado"] = 3.5 * 1.3
    processor.slas["trajeto_vazio"] = 2.9733 * 1.3
    
    return processor
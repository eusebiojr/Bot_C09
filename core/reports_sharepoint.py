# core/reports_sharepoint.py
"""
Módulo para gerenciar Reports diretamente no SharePoint.
Substitui dependência de arquivos locais por processamento em nuvem.
"""

import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Dict, Any, Optional
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


class SharePointReportsManager:
    """
    Gerencia sistema de Reports diretamente no SharePoint.
    Elimina dependência de arquivos locais.
    """
    
    def __init__(self, site_url: str, username: str, password: str):
        """
        Inicializa gerenciador de reports.
        
        Args:
            site_url: URL do site SharePoint
            username: Usuário SharePoint
            password: Senha SharePoint
        """
        self.site_url = site_url
        self.username = username
        self.password = password
        self._ctx = None
        
        # Caminhos SharePoint para reports
        self.base_docs = "/sites/Controleoperacional/Documentos Compartilhados/Bases de Dados"
        self.reports_path = f"{self.base_docs}/CREARE/Reports"
        self.reports_file = "base_dados_reports.xlsx"
    
    def _get_context(self) -> ClientContext:
        """Obtém contexto SharePoint."""
        if self._ctx is None:
            credentials = UserCredential(self.username, self.password)
            self._ctx = ClientContext(self.site_url).with_credentials(credentials)
        return self._ctx
    
    def _criar_pasta_reports(self) -> bool:
        """Cria pasta Reports no SharePoint se não existir."""
        try:
            ctx = self._get_context()
            
            # Verifica se pasta existe
            try:
                pasta = ctx.web.get_folder_by_server_relative_url(self.reports_path)
                ctx.load(pasta)
                ctx.execute_query()
                return True
            except:
                # Cria pasta
                pasta_pai = ctx.web.get_folder_by_server_relative_url(f"{self.base_docs}/CREARE")
                pasta_pai.folders.add("Reports")
                ctx.execute_query()
                print("✅ Pasta Reports criada no SharePoint")
                return True
                
        except Exception as e:
            print(f"❌ Erro ao criar pasta Reports: {e}")
            return False
    
    def carregar_arquivo_reports(self) -> Optional[pd.DataFrame]:
        """
        Carrega arquivo de reports do SharePoint.
        
        Returns:
            DataFrame com dados ou None se não existir
        """
        try:
            ctx = self._get_context()
            caminho_arquivo = f"{self.reports_path}/{self.reports_file}"
            
            # Baixa arquivo (método correto Office365)
            arquivo = ctx.web.get_file_by_server_relative_url(caminho_arquivo)
            
            # Cria buffer para receber o download
            download_buffer = BytesIO()
            arquivo.download(download_buffer)
            ctx.execute_query()
            
            # Volta para o início do buffer
            download_buffer.seek(0)
            
            # Carrega em DataFrame
            df = pd.read_excel(download_buffer, sheet_name="Resumo", engine="openpyxl")
            df['Data'] = pd.to_datetime(df['Data']).dt.date
            
            print(f"✅ Reports carregado: {len(df)} registros")
            return df
            
        except Exception as e:
            print(f"⚠️ Arquivo reports não encontrado no SharePoint: {e}")
            return None
    
    def carregar_candles_sharepoint(self, sheet_name: str) -> Optional[pd.DataFrame]:
        """
        Carrega dados de Candles do SharePoint.
        
        Args:
            sheet_name: Nome da aba (ex: "Candles", "Resumo por Hora")
            
        Returns:
            DataFrame com dados ou None se não existir
        """
        try:
            ctx = self._get_context()
            caminho_arquivo = f"{self.reports_path}/{self.reports_file}"
            
            # Baixa arquivo (método correto Office365)
            arquivo = ctx.web.get_file_by_server_relative_url(caminho_arquivo)
            
            # Cria buffer para receber o download
            download_buffer = BytesIO()
            arquivo.download(download_buffer)
            ctx.execute_query()
            
            # Volta para o início do buffer
            download_buffer.seek(0)
            
            # Carrega aba específica
            df = pd.read_excel(download_buffer, sheet_name=sheet_name, engine="openpyxl")
            
            # Converte colunas de data
            if sheet_name == "Candles" and "Data Evento" in df.columns:
                df["Data Evento"] = pd.to_datetime(df["Data Evento"], errors="coerce")
            elif sheet_name == "Resumo por Hora" and "Hora" in df.columns:
                df["Hora"] = pd.to_datetime(df["Hora"], errors="coerce")
            
            print(f"✅ {sheet_name} carregado: {len(df)} registros")
            return df
            
        except Exception as e:
            print(f"⚠️ {sheet_name} não encontrado: {e}")
            return pd.DataFrame()
    
    def salvar_arquivo_reports(self, df_resumo: pd.DataFrame, 
                              df_candles: Optional[pd.DataFrame] = None,
                              df_resumo_hora: Optional[pd.DataFrame] = None) -> bool:
        """
        Salva arquivo completo de reports no SharePoint.
        
        Args:
            df_resumo: DataFrame da aba Resumo
            df_candles: DataFrame da aba Candles (opcional)
            df_resumo_hora: DataFrame da aba Resumo por Hora (opcional)
            
        Returns:
            True se salvo com sucesso
        """
        try:
            # Cria pasta se necessário
            if not self._criar_pasta_reports():
                return False
            
            # Cria Excel em memória
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                # Aba principal
                df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
                
                # Abas opcionais
                if df_candles is not None and not df_candles.empty:
                    df_candles.to_excel(writer, sheet_name="Candles", index=False)
                
                if df_resumo_hora is not None and not df_resumo_hora.empty:
                    df_resumo_hora.to_excel(writer, sheet_name="Resumo por Hora", index=False)
            
            # Aplica formatação de tabelas
            buffer.seek(0)
            buffer_formatado = self._aplicar_formatacao_tabelas(buffer)
            
            # Upload para SharePoint
            ctx = self._get_context()
            caminho_arquivo = f"{self.reports_path}/{self.reports_file}"
            
            # Remove arquivo antigo
            try:
                arquivo_antigo = ctx.web.get_file_by_server_relative_url(caminho_arquivo)
                arquivo_antigo.delete_object()
                ctx.execute_query()
            except:
                pass  # Arquivo não existe, ok
            
            # Upload novo arquivo
            pasta_reports = ctx.web.get_folder_by_server_relative_url(self.reports_path)
            pasta_reports.upload_file(self.reports_file, buffer_formatado.read()).execute_query()
            
            print(f"✅ Reports salvo no SharePoint: {self.reports_file}")
            return True
            
        except Exception as e:
            print(f"❌ Erro ao salvar reports: {e}")
            return False
    
    def _aplicar_formatacao_tabelas(self, buffer: BytesIO) -> BytesIO:
        """Aplica formatação de tabelas ao Excel."""
        try:
            buffer.seek(0)
            wb = load_workbook(buffer)
            
            for aba_nome in wb.sheetnames:
                ws = wb[aba_nome]
                max_col = ws.max_column
                max_row = ws.max_row
                
                if max_row > 1 and max_col > 0:
                    # Calcula referência da tabela
                    col_letter_end = chr(ord('A') + max_col - 1)
                    ref = f"A1:{col_letter_end}{max_row}"
                    nome_tabela = f"Tabela_{aba_nome.replace(' ', '_')}"
                    
                    # Verifica se tabela já existe
                    tabelas_existentes = [t.displayName for t in ws.tables.values()]
                    if nome_tabela not in tabelas_existentes:
                        table = Table(displayName=nome_tabela, ref=ref)
                        style = TableStyleInfo(
                            name="TableStyleMedium9",
                            showFirstColumn=False,
                            showLastColumn=False,
                            showRowStripes=True,
                            showColumnStripes=False
                        )
                        table.tableStyleInfo = style
                        ws.add_table(table)
            
            # Salva resultado
            buffer_resultado = BytesIO()
            wb.save(buffer_resultado)
            buffer_resultado.seek(0)
            return buffer_resultado
            
        except Exception as e:
            print(f"⚠️ Erro na formatação, usando arquivo sem formatação: {e}")
            buffer.seek(0)
            return buffer
    
    def atualizar_resumo_diario(self, unidade: str, data: datetime.date, 
                               tpv_ac: float, dm_valor: float, total_veiculos: int) -> bool:
        """
        Atualiza resumo diário no arquivo reports.
        
        Args:
            unidade: Nome da unidade
            data: Data do registro
            tpv_ac: Valor TPV normalizado
            dm_valor: Horas de manutenção
            total_veiculos: Total de veículos da unidade
            
        Returns:
            True se atualizado com sucesso
        """
        try:
            # Carrega dados existentes
            df_existente = self.carregar_arquivo_reports()
            
            # Calcula DM percentual
            dm_percentual = 100 * (total_veiculos * 24 - dm_valor - 24 * 3) / (total_veiculos * 24)
            
            # Nova linha
            nova_linha = pd.DataFrame([{
                'Data': data,
                'TPV AC': tpv_ac / 24,  # Normaliza TPV
                'DM RRP': dm_percentual,
                'Unidade': unidade
            }])
            
            if df_existente is not None:
                # Verifica se data já existe
                if data in df_existente['Data'].values:
                    print(f"⚠️ Data {data} já existe no arquivo reports")
                    return True
                
                df_resultado = pd.concat([df_existente, nova_linha], ignore_index=True)
            else:
                df_resultado = nova_linha
            
            # Salva arquivo
            return self.salvar_arquivo_reports(df_resultado)
            
        except Exception as e:
            print(f"❌ Erro ao atualizar resumo diário: {e}")
            return False
    
    def atualizar_candles(self, df_eventos_novos: pd.DataFrame, 
                         df_resumo_novos: pd.DataFrame, poi: str, 
                         mes: int, ano: int) -> bool:
        """
        Atualiza dados de Candles no SharePoint.
        
        Args:
            df_eventos_novos: Novos eventos de entrada/saída
            df_resumo_novos: Novos dados de resumo por hora
            poi: POI dos dados
            mes: Mês de referência
            ano: Ano de referência
            
        Returns:
            True se atualizado com sucesso
        """
        try:
            # Carrega dados existentes
            df_candles_existente = self.carregar_candles_sharepoint("Candles")
            df_resumo_existente = self.carregar_candles_sharepoint("Resumo por Hora")
            df_resumo_geral = self.carregar_arquivo_reports()
            
            # Remove dados antigos do mesmo POI/mês/ano
            if not df_candles_existente.empty:
                df_candles_existente = df_candles_existente[
                    ~((df_candles_existente['Data Evento'].dt.month == mes) &
                      (df_candles_existente['Data Evento'].dt.year == ano) &
                      (df_candles_existente['POI'] == poi))
                ]
                df_candles_final = pd.concat([df_candles_existente, df_eventos_novos], ignore_index=True)
            else:
                df_candles_final = df_eventos_novos
            
            if not df_resumo_existente.empty:
                df_resumo_existente = df_resumo_existente[
                    ~((df_resumo_existente['Hora'].dt.month == mes) &
                      (df_resumo_existente['Hora'].dt.year == ano) &
                      (df_resumo_existente['POI'] == poi))
                ]
                df_resumo_hora_final = pd.concat([df_resumo_existente, df_resumo_novos], ignore_index=True)
            else:
                df_resumo_hora_final = df_resumo_novos
            
            # Usa DataFrame de resumo existente ou cria vazio
            if df_resumo_geral is None:
                df_resumo_geral = pd.DataFrame()
            
            # Salva arquivo completo
            return self.salvar_arquivo_reports(
                df_resumo=df_resumo_geral,
                df_candles=df_candles_final,
                df_resumo_hora=df_resumo_hora_final
            )
            
        except Exception as e:
            print(f"❌ Erro ao atualizar candles: {e}")
            return False


# Factory function
def criar_reports_manager(site_url: str = None, username: str = None, password: str = None) -> SharePointReportsManager:
    """Cria gerenciador de reports com configurações padrão."""
    import os
    from config.settings import ConstantesEspecificas
    
    if site_url is None:
        site_url = ConstantesEspecificas.SHAREPOINT_BASE_URL
    
    if username is None:
        username = os.getenv("SP_USER")
    
    if password is None:
        password = os.getenv("SP_PASSWORD")
    
    return SharePointReportsManager(site_url, username, password)